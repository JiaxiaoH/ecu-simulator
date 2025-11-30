# excel_io.py
import openpyxl
import yaml
from dataclasses import dataclass
from sessiontypes import SESSION_NAME_MAP_REVERSE
from security import SECURITY_NAME_MAP_REVERSE
RESPONSE_ID=0x18DAF1BA
PHYSICAL_ID=0x18DABAF1
FUNCTIONAL_ID=0x18DBBAF1
def normalize_sid_name(name: str) -> str | None:
    """
    get sid from name of worksheets
    suppot the following formats:
        "$SID10"
        "$SID_10"
        "$10"
        "SID10"
        "SID_10"
        "SID$10"
    """
    raw = name.replace("$", "").replace("_", "").replace("SID", "")
    raw = raw.upper()
    return raw
@dataclass
class ExcelTestCase:
    tcid: str
    row_index: int
    sheet_name: str
    request: str
    expected: str
    address: str
    session: str
    security: str

class ExcelTestCases:
    def __init__(self, workbook_path: str):
        self.workbook_path = workbook_path
        self.wb = openpyxl.load_workbook(workbook_path)
        self.cases: list[ExcelTestCase] = []
        self.sheet_header_map = {}     # {sheet_name: {header: column_index}}
        self.sheet_case_rows = {}      # {sheet_name: {tcid: row_index}}
    def append(self, object: ExcelTestCase, /) -> None:
        self.cases.append(object)
    def __iter__(self):
        return iter(self.cases)
    def __add__(self, other: 'ExcelTestCases') -> 'ExcelTestCases':
        self.cases.extend(other.cases)
        return self
    
    def _normalize_expected(self, sid, exp_raw):
        exp = exp_raw.strip()
        if exp.startswith("NRC$"):
            return f"7F {sid} {exp.replace('NRC$', '')}"
        if exp in ("ポジティブレスポンス", "Positive", "POS"):
            return "*"     # allow any positive response
        if exp in ("ノーレスポンス", "NO RESPONSE", "NoResponse"):
            return "NO RESPONSE"
        return exp
    
    def _address_filter(self, address: str)-> str:
        if address is None or address == "" or address == "-" or address.upper() == "物理" or address.upper() == "PHYSICAL":
            return hex(PHYSICAL_ID)
        if address == "機能" or address == "FUNCTIONAL":
            return hex(FUNCTIONAL_ID)

    def _adjust_hex_str(self, hex_str, length):
        parts = hex_str.split()
        count = len(parts)
        if count > length:
            parts = parts[:length]
        elif count < length:
            parts += ["00"] * (length - count)
        return " ".join(parts)

    def load(self):
        for ws in self.wb.worksheets:
            sid=normalize_sid_name(ws.title)
            if sid is None:
                continue
            headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
            self.sheet_header_map[ws.title] = headers
            self.sheet_case_rows[ws.title] = {}
            for row in ws.iter_rows(min_row=2):
                tcid = row[headers["TestCaseSpecID"]].value
                if tcid is None:
                    continue
                address_raw = row[headers["アドレス"]].value
                address = self._address_filter(address_raw)
                session_name = row[headers["Session"]].value
                if session_name is None:
                    continue
                session = SESSION_NAME_MAP_REVERSE[session_name]

                req_items = [sid]

                try:
                    subf_raw = (row[headers["sub-function"]].value or "").replace("$", "").strip()
                except KeyError:
                    subf_raw= None
                if subf_raw not in ("", "-", None):
                    subf_val = int(subf_raw, 16)
                    if row[headers["SPRMIB"]]is not None:
                        if row[headers["SPRMIB"]].value == 1:
                            subf_val |= 0x80
                            subf_raw=f"{subf_val:02X}"
                    req_items.append(subf_raw)
                
                data_raw = (row[headers.get("Dataパラメータ", None)]).value if "Dataパラメータ" in headers else ""
                if data_raw not in ("", None, "-"):
                    data_str = str(data_raw).replace("$", "").upper()
                    data_str2 = " ".join(data_str[i:i+2] for i in range(0, len(data_str), 2))
                    data=data_str2.split()
                    req_items.extend(data)

                req_str = " ".join(req_items)

                exp_raw = row[headers["期待値"]].value or ""
                expected = self._normalize_expected(sid, exp_raw)

                try:
                    security_raw = row[headers["security"]].value or ""
                    if security_raw not in ("", None, "-"):
                        security=SECURITY_NAME_MAP_REVERSE[security_raw]  
                    else:
                        security=None
                except KeyError:
                    security= None

                length_raw=row[headers["メッセージ長"]].value or ""
                if length_raw not in ("", None, "-"):
                    num_str = ''.join(ch for ch in length_raw if ch.isdigit())
                    length = int(num_str)
                    req_str=self._adjust_hex_str(req_str, length)
            
                self.cases.append(
                    ExcelTestCase(
                        tcid=str(tcid),
                        row_index=row[0].row,
                        sheet_name=ws.title,
                        request=req_str,
                        expected=expected,
                        address=address,
                        session=session,
                        security=security
                    )
                )
    def cases_to_yaml(self, output_path: str):
        yaml_cases = []
        for case in self.cases:
            yaml_case = {
                "TestCaseID": case.tcid,
                "Sheet": case.sheet_name,
                "Request": case.request,
                "Expected": case.expected,
                "Address": hex(FUNCTIONAL_ID) if case.address=="機能" else hex(PHYSICAL_ID), 
                "Session": case.session,  #
                "Row": case.row_index,
                "Security": case.security
            }
            yaml_cases.append(yaml_case)
        with open(output_path, 'w', encoding='utf-8') as f:
            yaml.dump(yaml_cases, f, allow_unicode=True)

    def write_back(self, tcid, actual, result):
        for case in self.cases:
            if case.tcid == tcid:
                ws = self.wb[case.sheet_name]
                headers = self.sheet_header_map[case.sheet_name]
                row = self.sheet_case_rows[case.sheet_name][case.tcid]
                ws.cell(row=row, column=len(headers)).value = result      
                ws.cell(row=row, column=len(headers)-1).value = actual    
                return

    def save(self):
        self.wb.save(self.workbook_path)

# if __name__ == "__main__":
#     load_excel_testcases("test_cases.xlsx")

